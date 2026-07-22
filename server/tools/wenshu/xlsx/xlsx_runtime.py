#!/usr/bin/env python3
"""WenShu spreadsheet runtime.

Consumes a JSON workbook spec and creates/modifies .xlsx files with openpyxl.
Calculations supplied by the caller remain Excel formulas so delivered models
stay linked, auditable and editable.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.comments import Comment

FORBIDDEN_FUNCTIONS = (
    "FILTER", "UNIQUE", "SORT", "SORTBY", "XLOOKUP", "XMATCH",
    "SEQUENCE", "LET", "LAMBDA", "RANDARRAY",
)


def emit(payload: dict[str, Any], code: int = 0) -> None:
    stream = sys.stdout if code == 0 else sys.stderr
    print(json.dumps(payload, ensure_ascii=False), file=stream)
    raise SystemExit(code)


def load_spec(path: str) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as fh:
        value = json.load(fh)
    if not isinstance(value, dict):
        raise ValueError("spec must be a JSON object")
    return value


def normalize_color(value: Any, default: str | None = None) -> str | None:
    if not isinstance(value, str) or not value.strip():
        return default
    color = value.strip().lstrip("#")
    if len(color) in (6, 8) and re.fullmatch(r"[0-9a-fA-F]+", color):
        return color.upper()
    return default


def make_side(spec: Any) -> Side:
    if not isinstance(spec, dict):
        return Side()
    return Side(style=spec.get("style"), color=normalize_color(spec.get("color")))


def apply_style(cell: Any, style: Any) -> None:
    if not isinstance(style, dict):
        return
    font_spec = style.get("font")
    if isinstance(font_spec, dict):
        cell.font = Font(
            name=font_spec.get("name"), sz=font_spec.get("size"), bold=font_spec.get("bold"),
            italic=font_spec.get("italic"), underline=font_spec.get("underline"),
            strike=font_spec.get("strike"), color=normalize_color(font_spec.get("color")),
        )
    fill_spec = style.get("fill")
    if isinstance(fill_spec, dict):
        color = normalize_color(fill_spec.get("color"), "FFFFFF")
        cell.fill = PatternFill(fill_type=fill_spec.get("type", "solid"), fgColor=color, bgColor=color)
    align_spec = style.get("alignment")
    if isinstance(align_spec, dict):
        cell.alignment = Alignment(
            horizontal=align_spec.get("horizontal"), vertical=align_spec.get("vertical"),
            wrap_text=align_spec.get("wrapText"), text_rotation=align_spec.get("textRotation", 0),
            shrink_to_fit=align_spec.get("shrinkToFit"), indent=align_spec.get("indent", 0),
        )
    border_spec = style.get("border")
    if isinstance(border_spec, dict):
        cell.border = Border(
            left=make_side(border_spec.get("left")), right=make_side(border_spec.get("right")),
            top=make_side(border_spec.get("top")), bottom=make_side(border_spec.get("bottom")),
        )
    if style.get("numberFormat") is not None:
        cell.number_format = str(style["numberFormat"])


def apply_cell(cell: Any, spec: dict[str, Any]) -> None:
    if "formula" in spec and spec.get("formula") is not None:
        formula = str(spec["formula"])
        cell.value = formula if formula.startswith("=") else f"={formula}"
    elif "value" in spec:
        cell.value = spec.get("value")
    apply_style(cell, spec.get("style"))
    if spec.get("numberFormat") is not None:
        cell.number_format = str(spec["numberFormat"])
    if isinstance(spec.get("comment"), dict):
        comment = spec["comment"]
        cell.comment = Comment(str(comment.get("text", "")), str(comment.get("author", "Mira")))
    if isinstance(spec.get("hyperlink"), str):
        cell.hyperlink = spec["hyperlink"]


def add_rows(ws: Any, rows: Any, start_row: int = 1, start_col: int = 1) -> None:
    if not isinstance(rows, list):
        return
    for row_offset, row in enumerate(rows):
        if not isinstance(row, list):
            continue
        for col_offset, value in enumerate(row):
            cell = ws.cell(row=start_row + row_offset, column=start_col + col_offset)
            if isinstance(value, dict):
                apply_cell(cell, value)
            else:
                cell.value = value


def add_cells(ws: Any, cells: Any) -> None:
    if not isinstance(cells, dict):
        return
    for address, value in cells.items():
        if not isinstance(address, str):
            continue
        cell = ws[address]
        if isinstance(value, dict):
            apply_cell(cell, value)
        else:
            cell.value = value


def apply_dimensions(ws: Any, spec: dict[str, Any]) -> None:
    columns = spec.get("columns")
    if isinstance(columns, dict):
        for column, value in columns.items():
            if not isinstance(column, str):
                continue
            dim = ws.column_dimensions[column]
            if isinstance(value, dict):
                if value.get("width") is not None: dim.width = float(value["width"])
                if value.get("hidden") is not None: dim.hidden = bool(value["hidden"])
            elif value is not None:
                dim.width = float(value)
    rows = spec.get("rowHeights")
    if isinstance(rows, dict):
        for row, height in rows.items():
            if height is not None:
                ws.row_dimensions[int(row)].height = float(height)


def add_conditional_formats(ws: Any, formats: Any) -> None:
    if not isinstance(formats, list):
        return
    for item in formats:
        if not isinstance(item, dict) or not isinstance(item.get("range"), str):
            continue
        kind = str(item.get("type", "colorScale"))
        target = item["range"]
        if kind == "colorScale":
            colors = item.get("colors") if isinstance(item.get("colors"), list) else ["F8696B", "FFEB84", "63BE7B"]
            colors = [normalize_color(v, "FFFFFF") for v in colors]
            if len(colors) >= 3:
                ws.conditional_formatting.add(target, ColorScaleRule(start_type="min", start_color=colors[0], mid_type="percentile", mid_value=50, mid_color=colors[1], end_type="max", end_color=colors[2]))
        elif kind == "cellIs":
            formula = item.get("formula") if isinstance(item.get("formula"), list) else [str(item.get("formula", "0"))]
            fill = PatternFill(fill_type="solid", fgColor=normalize_color(item.get("fill"), "FFF2CC"))
            ws.conditional_formatting.add(target, CellIsRule(operator=str(item.get("operator", "greaterThan")), formula=formula, fill=fill))
        elif kind == "formula":
            formula = item.get("formula") if isinstance(item.get("formula"), list) else [str(item.get("formula", "TRUE"))]
            fill = PatternFill(fill_type="solid", fgColor=normalize_color(item.get("fill"), "FFF2CC"))
            ws.conditional_formatting.add(target, FormulaRule(formula=formula, fill=fill))


def add_chart(ws: Any, item: dict[str, Any]) -> None:
    kind = str(item.get("type", "column")).lower()
    chart: Any
    if kind == "line":
        chart = LineChart()
    elif kind == "pie":
        chart = PieChart()
    else:
        chart = BarChart()
        chart.type = "bar" if kind == "bar" else "col"
        chart.style = 10
    chart.title = str(item.get("title", "")) or None
    chart.height = float(item.get("height", 7.5))
    chart.width = float(item.get("width", 12))
    if item.get("yTitle"): chart.y_axis.title = str(item["yTitle"])
    if item.get("xTitle") and hasattr(chart, "x_axis"): chart.x_axis.title = str(item["xTitle"])
    data_spec = item.get("data") if isinstance(item.get("data"), dict) else {}
    if not data_spec:
        return
    data = Reference(ws, min_col=int(data_spec.get("minCol", 2)), max_col=int(data_spec.get("maxCol", 2)), min_row=int(data_spec.get("minRow", 1)), max_row=int(data_spec.get("maxRow", ws.max_row)))
    categories = None
    cats = item.get("categories") if isinstance(item.get("categories"), dict) else {}
    if cats:
        categories = Reference(ws, min_col=int(cats.get("minCol", 1)), max_col=int(cats.get("maxCol", 1)), min_row=int(cats.get("minRow", 2)), max_row=int(cats.get("maxRow", ws.max_row)))
    chart.add_data(data, titles_from_data=bool(data_spec.get("titlesFromData", True)))
    if categories is not None: chart.set_categories(categories)
    ws.add_chart(chart, str(item.get("anchor", "H2")))


def apply_sheet_spec(ws: Any, spec: dict[str, Any]) -> None:
    if "rows" in spec:
        add_rows(ws, spec.get("rows"), int(spec.get("startRow", 1)), int(spec.get("startColumn", 1)))
    add_cells(ws, spec.get("cells"))
    for merge in spec.get("merges", []) or []:
        if isinstance(merge, str): ws.merge_cells(merge)
    apply_dimensions(ws, spec)
    if isinstance(spec.get("freezePanes"), str): ws.freeze_panes = spec["freezePanes"]
    if spec.get("showGridLines") is not None: ws.sheet_view.showGridLines = bool(spec["showGridLines"])
    if isinstance(spec.get("tabColor"), str): ws.sheet_properties.tabColor = normalize_color(spec["tabColor"])
    add_conditional_formats(ws, spec.get("conditionalFormats"))
    for chart_spec in spec.get("charts", []) or []:
        if isinstance(chart_spec, dict): add_chart(ws, chart_spec)


def add_sources_sheet(wb: Any, sources: Any) -> None:
    if not isinstance(sources, list) or not sources:
        return
    if "Sources" in wb.sheetnames:
        ws = wb["Sources"]
        if ws.max_row == 1 and ws["A1"].value is None: ws.delete_rows(1)
    else:
        ws = wb.create_sheet("Sources")
    if ws.max_row == 1 and ws["A1"].value is None:
        ws.append(["Source Name", "Source URL", "Notes"])
    elif ws.max_row == 0:
        ws.append(["Source Name", "Source URL", "Notes"])
    for item in sources:
        if isinstance(item, dict):
            ws.append([item.get("name", ""), item.get("url", ""), item.get("notes", "")])
    for cell in ws[1]: cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 50


def apply_named_ranges(wb: Any, named_ranges: Any) -> None:
    if not isinstance(named_ranges, list):
        return
    from openpyxl.workbook.defined_name import DefinedName
    for item in named_ranges:
        if not isinstance(item, dict) or not item.get("name") or not item.get("ref"):
            continue
        name = str(item["name"])
        try:
            wb.defined_names.delete(name)
        except Exception:
            pass
        wb.defined_names.add(DefinedName(name, attr_text=str(item["ref"])))


def apply_workbook_spec(wb: Any, spec: dict[str, Any], create: bool) -> None:
    meta = spec.get("metadata") if isinstance(spec.get("metadata"), dict) else {}
    if meta:
        if meta.get("creator") is not None: wb.creator = str(meta["creator"])
        if meta.get("title") is not None: wb.title = str(meta["title"])
        if meta.get("subject") is not None: wb.subject = str(meta["subject"])
        if meta.get("description") is not None: wb.description = str(meta["description"])
    remove_sheets = spec.get("removeSheets")
    if isinstance(remove_sheets, list):
        for name in remove_sheets:
            if isinstance(name, str) and name in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb[name])
    first_created = create
    for sheet_spec in spec.get("sheets", []) or []:
        if not isinstance(sheet_spec, dict):
            continue
        name = str(sheet_spec.get("name", "Sheet"))[:31] or "Sheet"
        if name in wb.sheetnames:
            ws = wb[name]
        elif first_created and wb.sheetnames == ["Sheet"] and wb["Sheet"]["A1"].value is None:
            ws = wb["Sheet"]
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        first_created = False
        apply_sheet_spec(ws, sheet_spec)
    add_sources_sheet(wb, spec.get("sources", []) or [])
    apply_named_ranges(wb, spec.get("namedRanges"))
    if wb.calculation is not None:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True


def create_workbook(spec: dict[str, Any], output: str) -> dict[str, Any]:
    wb = Workbook()
    apply_workbook_spec(wb, spec, create=True)
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"output": output, "sheets": wb.sheetnames}


def modify_workbook(input_path: str, spec: dict[str, Any], output: str) -> dict[str, Any]:
    wb = load_workbook(input_path)
    apply_workbook_spec(wb, spec, create=False)
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {"output": output, "sheets": wb.sheetnames}


def inspect_workbook(input_path: str) -> dict[str, Any]:
    wb = load_workbook(input_path, data_only=False, read_only=False)
    sheets = []
    formula_count = 0
    for ws in wb.worksheets:
        local_formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    local_formulas += 1
        formula_count += local_formulas
        sheets.append({
            "name": ws.title, "rows": ws.max_row, "columns": ws.max_column,
            "formulas": local_formulas, "mergedRanges": len(ws.merged_cells.ranges),
            "charts": len(ws._charts),
        })
    return {"file": input_path, "sheetCount": len(sheets), "formulaCount": formula_count, "sheets": sheets}


def verify_workbook(input_path: str) -> dict[str, Any]:
    wb = load_workbook(input_path, data_only=False)
    issues: list[dict[str, Any]] = []
    formulas = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                    upper = value.upper()
                    for fn in FORBIDDEN_FUNCTIONS:
                        if re.search(rf"\b{re.escape(fn)}\s*\(", upper):
                            issues.append({"type": "forbidden_function", "location": f"{ws.title}!{cell.coordinate}", "function": fn})
                    if re.search(r"MATCH\s*\(\s*TRUE\s*\(", upper):
                        issues.append({"type": "implicit_array_formula", "location": f"{ws.title}!{cell.coordinate}"})
    try:
        data_wb = load_workbook(input_path, data_only=True)
        for ws in data_wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("#"):
                        issues.append({"type": "formula_error", "location": f"{ws.title}!{cell.coordinate}", "value": cell.value})
    except Exception as exc:
        issues.append({"type": "cached_value_check_failed", "message": str(exc)})
    blocking = [item for item in issues if item["type"] not in ("cached_value_check_failed", "forbidden_function", "implicit_array_formula")]
    return {"file": input_path, "formulas": formulas, "issues": issues, "ok": len(blocking) == 0}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("operation", choices=["create", "modify", "inspect", "verify"])
    parser.add_argument("--input")
    parser.add_argument("--output")
    parser.add_argument("--spec")
    args = parser.parse_args()
    try:
        if args.operation == "create":
            if not args.output or not args.spec: raise ValueError("create requires --output and --spec")
            result = create_workbook(load_spec(args.spec), args.output)
        elif args.operation == "modify":
            if not args.input or not args.output or not args.spec: raise ValueError("modify requires --input, --output and --spec")
            result = modify_workbook(args.input, load_spec(args.spec), args.output)
        elif args.operation == "inspect":
            if not args.input: raise ValueError("inspect requires --input")
            result = inspect_workbook(args.input)
        else:
            if not args.input: raise ValueError("verify requires --input")
            result = verify_workbook(args.input)
        emit({"status": "success", "data": result})
    except Exception as exc:
        emit({"status": "error", "error": type(exc).__name__, "message": str(exc)}, 1)


if __name__ == "__main__":
    main()
